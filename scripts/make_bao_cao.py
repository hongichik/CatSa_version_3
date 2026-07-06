#!/usr/bin/env python3
"""Tổng hợp báo cáo CatSA vs CORE (demo2 + test_all) -> Excel.

Mẫu tham chiếu: test_all/scripts/make_v1_core_full_report.py
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
LOG_RETAIL = REPO / "Log" / "retailrocket"
LOG_DIGI = REPO / "Log" / "diginetica"
TEST_ALL_JSONL = Path("/home/hongnguyen/test_all/Log/_queue/v1_core_eval/results-2026-07-03-01-53-23.jsonl")

GREEN_F = PatternFill("solid", fgColor="C6EFCE")
GREEN_T = Font(color="006100", bold=True)
RED_F = PatternFill("solid", fgColor="FFC7CE")
RED_T = Font(color="9C0006", bold=True)
HEAD_F = PatternFill("solid", fgColor="305496")
HEAD_T = Font(color="FFFFFF", bold=True)
GRAY_F = PatternFill("solid", fgColor="F2F2F2")
SUB_F = PatternFill("solid", fgColor="D9E1F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- Ghi chú chuẩn cho tên dataset / config ---
DATASET_NOTES = {
    "retailrocket": "RR full (item_hon_5): giữ mọi phiên, prefix≤50, có taxonomy",
    "retailrocket_item_hon_5": "RR full: support≥5, keep session, max_prefix=50",
    "retailrocket_2_5": "Chỉ phiên độ dài 2–5 click (session_length_mode=filter)",
    "retailrocket_2_7": "Chỉ phiên độ dài 2–7 click",
    "retailrocket_3_6": "Chỉ phiên độ dài 3–6 click",
    "retailrocket_4_7": "Chỉ phiên độ dài 4–7 click",
    "retailrocket_1_4": "Subset RR: max_seq/prefix = 4 (test_all)",
    "retailrocket_3_7": "Subset RR: max_seq/prefix = 7 (test_all)",
    "retailrocket_category": "RR: chỉ item có category (require_item_category=true)",
    "retailrocket_same_leaf": "Subset test: target cùng leaf category với item cuối session",
    "retailrocket_cold": "Subset test: cold-start / item hiếm",
    "diginetica": "Diginetica: category phẳng, use_taxonomy=false",
}

AUGMENT_NOTES = {
    "v1": "Augment: same + sibling + hybrid (đủ 3 chiến lược); use_cl=true",
    "v2": "Augment: chỉ same-leaf; use_cl=true",
    "v3": "Augment: chỉ sibling (RR) / hybrid (Diginetica); use_cl=true",
    "v4": "Augment: chỉ hybrid; use_cl=true",
    "a2": "A2 — Module 1 only: use_cl=false, chỉ L_rec",
    "full": "CatSA full — Module 1+2+3: use_cl=true + InfoNCE",
    "tune_a2": "Tune A2 Diginetica: num_layers=1, dropout=0.2, batch=128, no CL",
    "tune_cl": "Tune CL nhẹ: lambda_cl=0.01, batch=128, same augment",
    "tune_batch256": "Tune A2 + batch_size=256",
    "tune_batch2048": "Tune A2 + batch_size=2048 (align CORE)",
    "core_trm": "CORE-trm: Transformer + cosine/temperature, batch=256/2048",
    "core_ave": "CORE-ave: average pooling encoder",
}

CATSA_V1_SPEC = [
    ("Tên model", "CatSA (RGCN + category graph)", "CORE-trm (Transformer)"),
    ("Backbone", "Hetero graph + RGCN (SAGEConv)", "TransNet + attention pooling"),
    ("Số lớp encoder", "num_layers=1–2 (tune: 1)", "n_layers=2"),
    ("Embedding dim", "100", "100"),
    ("Taxonomy", "RR: có; Diginetica: không", "Không"),
    ("Augment / CL", "same/sibling/hybrid + InfoNCE (tùy config)", "Không"),
    ("Scoring", "Dot-product z·item_emb", "L2-norm cosine / temperature"),
    ("max_prefix_length", "50", "50"),
    ("batch_size", "100–2048", "256 (demo2) / 2048 (test_all)"),
    ("Metric", "MRR@20 (full-ranking test)", "MRR@20 (full-ranking test)"),
]


def _parse_test_line(line: str) -> dict[str, float]:
    m = re.search(
        r"hr@10=([\d.]+).*mrr@10=([\d.]+).*hr@20=([\d.]+).*mrr@20=([\d.]+)",
        line,
    )
    if not m:
        return {}
    h10, m10, h20, m20 = map(float, m.groups())
    return {
        "HR@10": round(h10 * 100, 2),
        "MRR@10": round(m10 * 100, 2),
        "HR@20": round(h20 * 100, 2),
        "MRR@20": round(m20 * 100, 2),
    }


def _parse_core_test(line: str) -> dict[str, float]:
    m = re.search(r"test result:\s*(\{.*\})", line)
    if not m:
        # fallback: dòng valid/test dạng recall@10 : 0.5763 ...
        pairs = re.findall(
            r"(recall|mrr)@(\d+)\s*:\s*([\d.]+)", line, re.I
        )
        if not pairs:
            return {}
        d: dict[str, float] = {}
        for kind, k, v in pairs:
            key = f"{'recall' if kind.lower() == 'recall' else 'mrr'}@{k}"
            d[key] = float(v)
        return {
            "HR@10": round(d.get("recall@10", 0) * 100, 2),
            "MRR@10": round(d.get("mrr@10", 0) * 100, 2),
            "HR@20": round(d.get("recall@20", 0) * 100, 2),
            "MRR@20": round(d.get("mrr@20", 0) * 100, 2),
        }
    raw = m.group(1).replace("np.float64(", "").replace(")", "")
    pairs = re.findall(r"['\"]?([\w@]+)['\"]?\s*:\s*([\d.]+)", raw)
    d = {k: float(v) for k, v in pairs}
    return {
        "HR@10": round(d.get("recall@10", 0) * 100, 2),
        "MRR@10": round(d.get("mrr@10", 0) * 100, 2),
        "HR@20": round(d.get("recall@20", 0) * 100, 2),
        "MRR@20": round(d.get("mrr@20", 0) * 100, 2),
    }


def _load_demo2_rows() -> list[dict]:
    rows: list[dict] = []
    mapping = [
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_v1.log", "v1", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_v2.log", "v2", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_v3.log", "v3", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_v4.log", "v4", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_tune_a2.log", "tune_a2", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_tune_cl.log", "tune_cl", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_tune_batch256.log", "tune_batch256", "demo2"),
        ("diginetica", "CatSA", LOG_DIGI, "diginetica_tune_batch2048.log", "tune_batch2048", "demo2"),
        ("diginetica", "CORE", LOG_DIGI, "core_trm_diginetica.log", "core_trm", "demo2"),
        ("diginetica", "CORE", LOG_DIGI, "core_ave_diginetica.log", "core_ave", "demo2"),
        ("retailrocket_item_hon_5", "CatSA", LOG_RETAIL, "CatSA_v1.log", "v1", "demo2"),
        ("retailrocket_item_hon_5", "CatSA", LOG_RETAIL, "CatSA_v2.log", "v2", "demo2"),
        ("retailrocket_item_hon_5", "CatSA", LOG_RETAIL, "CatSA_v3.log", "v3", "demo2"),
        ("retailrocket_item_hon_5", "CatSA", LOG_RETAIL, "CatSA_v4.log", "v4", "demo2"),
        ("retailrocket_2_5", "CatSA", LOG_RETAIL, "CatSA_2_5.log", "v1", "demo2"),
        ("retailrocket_2_7", "CatSA", LOG_RETAIL, "CatSA_2_7.log", "v1", "demo2"),
        ("retailrocket_3_6", "CatSA", LOG_RETAIL, "CatSA_3_6.log", "v1", "demo2"),
        ("retailrocket_4_7", "CatSA", LOG_RETAIL, "CatSA_4_7.log", "v1", "demo2"),
        ("retailrocket_category", "CatSA", LOG_RETAIL, "CatSA_category.log", "v1", "demo2"),
        ("retailrocket_item_hon_5", "CORE", LOG_RETAIL, "core_trm_retailrocket.log", "core_trm", "demo2"),
        ("retailrocket_1_4", "CORE", LOG_RETAIL, "core_trm_retailrocket_1_4.log", "core_trm", "demo2"),
        ("retailrocket_3_7", "CORE", LOG_RETAIL, "core_trm_retailrocket_3_7.log", "core_trm", "demo2"),
        ("retailrocket_same_leaf", "CORE", LOG_RETAIL, "core_trm_retailrocket_same_leaf.log", "core_trm", "demo2"),
        ("retailrocket_cold", "CORE", LOG_RETAIL, "core_trm_retailrocket_cold.log", "core_trm", "demo2"),
    ]
    for ds, model, log_dir, logname, cfg, src in mapping:
        path = log_dir / logname
        if not path.is_file():
            continue
        text = path.read_text()
        if "KẾT QUẢ TEST" in text:
            line = [l for l in text.splitlines() if "KẾT QUẢ TEST" in l][-1]
            metrics = _parse_test_line(line)
        else:
            line = [l for l in text.splitlines() if "test result:" in l][-1]
            metrics = _parse_core_test(line)
        if not metrics:
            continue
        note_parts = [DATASET_NOTES.get(ds, ds)]
        if model == "CatSA":
            note_parts.append(AUGMENT_NOTES.get(cfg, cfg))
        else:
            note_parts.append(AUGMENT_NOTES.get(cfg, cfg))
        rows.append({
            "dataset": ds,
            "model": model,
            "config": cfg,
            "mode": cfg if cfg in ("a2", "full") else "—",
            "source": src,
            "note": "; ".join(note_parts),
            **metrics,
        })
    return rows


def _load_jsonl_cmp() -> list[dict]:
    if not TEST_ALL_JSONL.is_file():
        return []
    out = []
    for line in TEST_ALL_JSONL.read_text().splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        o = rec["overall"]
        ds = o["dataset"]
        mode = o.get("mode", "a2")
        ds_note = DATASET_NOTES.get(ds, ds)
        mode_note = AUGMENT_NOTES.get(mode, mode)
        out.append({
            "dataset": ds,
            "model": "CatSA vs CORE",
            "config": f"v1-{mode}",
            "mode": mode,
            "source": "test_all",
            "note": f"{ds_note}; CatSA v1 {mode_note}; CORE-trm batch=2048",
            "n": o["n"],
            "catsa_mrr20": round(o["catsa"]["MRR@20"], 2),
            "core_mrr20": round(o["core"]["MRR@20"], 2),
            "delta_mrr20": round(o["delta_mrr@20"], 2),
            "catsa_hit20": round(o["catsa"]["Recall@20"], 2),
            "core_hit20": round(o["core"]["Recall@20"], 2),
            "delta_hit20": round(o["delta_hit@20"], 2),
        })
    return out


def _style_header(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill, cell.font, cell.alignment, cell.border = HEAD_F, HEAD_T, CENTER, BORDER


def _cmp_text(c: float, k: float) -> tuple[str, bool | None]:
    gap = round(abs(c - k), 2)
    if c < k:
        return f"{c:.2f} - {gap:.2f} | {k:.2f}", False
    if c > k:
        return f"{c:.2f} + {gap:.2f} | {k:.2f}", True
    return f"{c:.2f} = | {k:.2f}", None


def _style_cmp(cell, win: bool | None) -> None:
    cell.alignment = CENTER
    cell.border = BORDER
    if win is False:
        cell.fill, cell.font = RED_F, RED_T
    elif win is True:
        cell.fill, cell.font = GREEN_F, GREEN_T


def write_report(out: Path) -> None:
    demo2 = _load_demo2_rows()
    jsonl = _load_jsonl_cmp()
    wb = Workbook()

    # Sheet 1: Kiến trúc
    ws = wb.active
    ws.title = "Kiến trúc"
    _style_header(ws, ["Thông số", "CatSA", "CORE-trm"])
    for r, (a, b, c) in enumerate(CATSA_V1_SPEC, 2):
        ws.cell(r, 1, a).font = Font(bold=True)
        ws.cell(r, 1).fill = GRAY_F
        ws.cell(r, 2, b)
        ws.cell(r, 3, c)
        for col in range(1, 4):
            ws.cell(r, col).alignment = LEFT
            ws.cell(r, col).border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40

    # Sheet 2: Tổng hợp demo2 (có Ghi chú)
    ws2 = wb.create_sheet("Kết quả demo2")
    h2 = [
        "Dataset", "Model", "Config/Run", "MRR@10", "MRR@20", "HR@10", "HR@20",
        "Nguồn", "Ghi chú",
    ]
    _style_header(ws2, h2)
    for r, row in enumerate(demo2, 2):
        ws2.cell(r, 1, row["dataset"])
        ws2.cell(r, 2, row["model"])
        ws2.cell(r, 3, row["config"])
        ws2.cell(r, 4, row["MRR@10"])
        ws2.cell(r, 5, row["MRR@20"])
        ws2.cell(r, 6, row["HR@10"])
        ws2.cell(r, 7, row["HR@20"])
        ws2.cell(r, 8, row["source"])
        ws2.cell(r, 9, row["note"])
        ws2.cell(r, 9).alignment = LEFT
        for c in range(1, 9):
            ws2.cell(r, c).alignment = CENTER
        for c in range(1, 10):
            ws2.cell(r, c).border = BORDER
        if row["model"] == "CORE":
            for c in range(1, 8):
                ws2.cell(r, c).fill = SUB_F
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["I"].width = 62
    for col in range(2, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    ws2.freeze_panes = "A2"

    # Sheet 3: CatSA vs CORE (test_all v1 eval)
    ws3 = wb.create_sheet("CatSA vs CORE")
    h3 = [
        "Dataset", "Mode", "N test",
        "MRR@20 (CatSA-Δ|CORE)", "Hit@20 (CatSA-Δ|CORE)",
        "Δ MRR", "Δ Hit", "Ghi chú",
    ]
    _style_header(ws3, h3)
    for r, row in enumerate(jsonl, 2):
        ws3.cell(r, 1, row["dataset"])
        ws3.cell(r, 2, row["mode"])
        ws3.cell(r, 3, row["n"])
        t_mrr, w_mrr = _cmp_text(row["catsa_mrr20"], row["core_mrr20"])
        t_hit, w_hit = _cmp_text(row["catsa_hit20"], row["core_hit20"])
        _style_cmp(ws3.cell(r, 4, t_mrr), w_mrr)
        _style_cmp(ws3.cell(r, 5, t_hit), w_hit)
        ws3.cell(r, 6, row["delta_mrr20"])
        ws3.cell(r, 7, row["delta_hit20"])
        ws3.cell(r, 8, row["note"])
        ws3.cell(r, 8).alignment = LEFT
        for c in (1, 2, 3, 6, 7):
            ws3.cell(r, c).alignment = CENTER
            ws3.cell(r, c).border = BORDER
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["H"].width = 58
    ws3.freeze_panes = "A2"

    # Sheet 4: Chú thích tên
    ws4 = wb.create_sheet("Chú thích tên")
    legend = [
        ("Tên", "Ý nghĩa"),
        ("2_5", "Dataset preprocess: chỉ giữ phiên có 2–5 click"),
        ("2_7", "Dataset preprocess: chỉ giữ phiên có 2–7 click"),
        ("3_6", "Dataset preprocess: chỉ giữ phiên có 3–6 click"),
        ("4_7", "Dataset preprocess: chỉ giữ phiên có 4–7 click"),
        ("item_hon_5 / full", "Toàn bộ phiên (support≥5), cắt prefix≤50 lúc train"),
        ("v1 (CatSA)", "Augment: same + sibling + hybrid; CL bật"),
        ("v2", "Augment: chỉ same-leaf"),
        ("v3 (RR)", "Augment: chỉ sibling | (Diginetica: hybrid)"),
        ("v4", "Augment: chỉ hybrid (same-leaf rồi random crop)"),
        ("a2", "CatSA v1 không CL — chỉ Module 1 (L_rec)"),
        ("full", "CatSA v1 đủ Module 1+2+3 (CL + augment)"),
        ("same_leaf", "Subset đánh giá: target cùng category lá với click cuối"),
        ("cold", "Subset cold-start / item hiếm trong test"),
        ("1_4 / 3_7", "Subset RR giới hạn độ dài sequence 4 hoặc 7"),
        ("tune_*", "Thí nghiệm tune CatSA trên Diginetica (xem config)"),
    ]
    for r, (k, v) in enumerate(legend, 1):
        ws4.cell(r, 1, k).font = Font(bold=(r == 1))
        ws4.cell(r, 2, v)
        ws4.cell(r, 1).border = ws4.cell(r, 2).border = BORDER
        ws4.cell(r, 2).alignment = LEFT
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 65

    # Sheet 5: Kết luận
    ws5 = wb.create_sheet("Kết luận")
    digi_catsa = [r for r in demo2 if r["dataset"] == "diginetica" and r["model"] == "CatSA"]
    best_digi = max(digi_catsa, key=lambda x: x["MRR@20"]) if digi_catsa else None
    core_digi = next((r for r in demo2 if r["config"] == "core_trm" and r["dataset"] == "diginetica"), None)
    lines = [
        f"Báo cáo tổng hợp CatSA vs CORE — sinh {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "Diginetica (demo2):",
    ]
    if best_digi and core_digi:
        lines.append(
            f"• CatSA tốt nhất: {best_digi['config']} MRR@20={best_digi['MRR@20']}% "
            f"(batch2048={best_digi['config']=='tune_batch2048'})"
        )
        lines.append(f"• CORE-trm: MRR@20={core_digi['MRR@20']}% — thua ~{core_digi['MRR@20']-best_digi['MRR@20']:.2f}pp")
    lines.extend([
        "",
        "RetailRocket session length (demo2, CatSA v1 full augment):",
        "• 2_5 (phiên 2–5): MRR@20 cao nhất trong ablation độ dài (~36.7%)",
        "• Full item_hon_5 baseline v1: ~28.1% (2 lớp GNN + CL)",
        "",
        "test_all CatSA v1 vs CORE (a2/full):",
        "• Full RR: gần hòa với CatSA v7 (~38.6% vs ~38.8% trong report cũ)",
        "• same_leaf segment: CatSA có thể thắng CORE về MRR",
        "• CatSA v1 a2 thua CORE ~10pp trên full RR test",
        "",
        "Ghi chú: metric HR@K = Recall@K; đơn vị %.",
    ])
    for i, line in enumerate(lines, 1):
        ws5.cell(i, 1, line)
        ws5.cell(i, 1).alignment = LEFT
    ws5.column_dimensions["A"].width = 95

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> None:
    ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    out = REPO / "Log" / f"bao_cao_catsa_core_{ts}.xlsx"
    write_report(out)
    print(f"Đã ghi: {out}")


if __name__ == "__main__":
    main()
